"""
DL (PBertKla) 16샘플 (new_data_v1.csv) 결과 정리
- Excel: data1/2/3 각각의 fold별 + 5-fold 평균 metric (Precision, Recall, F1, Acc, AUC-ROC, AUC-PRC)
- ROC: 16샘플이므로 IJMS 'star' (step) 스타일 사용
"""
import os
import numpy as np
import pandas as pd
import matplotlib as mpl
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
from sklearn.metrics import (roc_curve, roc_auc_score, average_precision_score,
                              precision_score, recall_score, f1_score,
                              accuracy_score)

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── IJMS 폰트 ──
_PREFERRED = ["Arial", "Helvetica", "Liberation Sans", "DejaVu Sans"]
mpl.rcParams["font.family"]     = "sans-serif"
mpl.rcParams["font.sans-serif"] = _PREFERRED
mpl.rcParams["pdf.fonttype"]    = 42
mpl.rcParams["ps.fonttype"]     = 42
mpl.rcParams["svg.fonttype"]    = "none"

# ── 경로 ──
BASE        = "/home/work/LNP_TEST/git_tools/PBertKla_v2"
PRED_ROOT   = f"{BASE}/results/Inference/DL_preds_5fold"   # data{N}_newdata/fold_k/predictions.npy
DATA_CSV    = f"{BASE}/Data/4_infer_new_data/new_data_v1.csv"

OUT_XLSX    = f"{BASE}/results/DL_newdata_metrics.xlsx"
OUT_ROC_DIR = f"{BASE}/results/DL_newdata_roc"

DATASETS    = ["data1", "data2", "data3"]
THRESH      = 0.5
FIGSIZE_IN  = 4.0
DPI         = 300

CURVE_COLOR = "#0072B2"
AVG_COLOR   = "#D55E00"
FOLD_COLORS = ["#0072B2", "#E69F00", "#009E73", "#CC79A7", "#56B4E9"]


# ───────── ROC 그리기 (step, IJMS star) ─────────
def _save_fig(fig, save_dir, name, dpi=DPI):
    os.makedirs(save_dir, exist_ok=True)
    fig.savefig(f"{save_dir}/{name}.png", dpi=dpi, facecolor="white", bbox_inches="tight")
    fig.savefig(f"{save_dir}/{name}.pdf",         facecolor="white", bbox_inches="tight")
    fig.savefig(f"{save_dir}/{name}.tif", dpi=dpi, format="tiff",
                facecolor="white", bbox_inches="tight",
                pil_kwargs={"compression": "tiff_lzw"})


def plot_roc_step(y_true, y_pred, title, save_dir, fname, color=CURVE_COLOR):
    """16 샘플처럼 작은 셋용 — step 함수 (보간 없음)"""
    fpr, tpr, _ = roc_curve(y_true, y_pred)
    auc = roc_auc_score(y_true, y_pred)

    fig, ax = plt.subplots(figsize=(FIGSIZE_IN, FIGSIZE_IN), facecolor="white")
    ax.step(fpr, tpr, where="post", linewidth=1.8, color=color,
            label=f"ROC (AUC = {auc:.3f})")
    ax.plot([0, 1], [0, 1], "--", linewidth=1.0, color="grey", label="Random")

    ax.set_xlim(0.0, 1.0); ax.set_ylim(0.0, 1.02)
    ax.set_xlabel("False Positive Rate", fontsize=10)
    ax.set_ylabel("True Positive Rate",  fontsize=10)
    ax.set_title(title, fontsize=11)
    ax.tick_params(axis="both", labelsize=9)
    ax.legend(loc="lower right", fontsize=9, framealpha=0.95)
    ax.grid(True, linestyle=":", linewidth=0.5, alpha=0.6)
    ax.set_aspect("equal")
    for spine in ax.spines.values():
        spine.set_linewidth(0.8)
    fig.tight_layout()
    _save_fig(fig, save_dir, fname)
    plt.close(fig)
    return auc


def plot_roc_overlay_step(folds_data, avg_data, title, save_dir, fname):
    fig, ax = plt.subplots(figsize=(FIGSIZE_IN, FIGSIZE_IN), facecolor="white")
    for label, y_true, y_pred, color in folds_data:
        fpr, tpr, _ = roc_curve(y_true, y_pred)
        auc = roc_auc_score(y_true, y_pred)
        ax.step(fpr, tpr, where="post", linewidth=1.0, color=color, alpha=0.7,
                label=f"{label} (AUC = {auc:.3f})")
    label, y_true, y_pred, color = avg_data
    fpr, tpr, _ = roc_curve(y_true, y_pred)
    auc = roc_auc_score(y_true, y_pred)
    ax.step(fpr, tpr, where="post", linewidth=2.4, color=color,
            label=f"{label} (AUC = {auc:.3f})")
    ax.plot([0, 1], [0, 1], "--", linewidth=1.0, color="grey", label="Random")

    ax.set_xlim(0.0, 1.0); ax.set_ylim(0.0, 1.02)
    ax.set_xlabel("False Positive Rate", fontsize=10)
    ax.set_ylabel("True Positive Rate",  fontsize=10)
    ax.set_title(title, fontsize=11)
    ax.tick_params(axis="both", labelsize=9)
    ax.legend(loc="lower right", fontsize=8, framealpha=0.95)
    ax.grid(True, linestyle=":", linewidth=0.5, alpha=0.6)
    ax.set_aspect("equal")
    for spine in ax.spines.values():
        spine.set_linewidth(0.8)
    fig.tight_layout()
    _save_fig(fig, save_dir, fname)
    plt.close(fig)


# ───────── 메트릭 계산 ─────────
def compute_metrics(y_true, y_pred):
    y_bin = (y_pred >= THRESH).astype(int)
    return {
        "Accuracy":  accuracy_score(y_true, y_bin),
        "Precision": precision_score(y_true, y_bin, zero_division=0),
        "Recall":    recall_score(y_true, y_bin, zero_division=0),
        "F1":        f1_score(y_true, y_bin, zero_division=0),
        "AUC-ROC":   roc_auc_score(y_true, y_pred),
        "AUC-PRC":   average_precision_score(y_true, y_pred),
    }


# ───────── 메인 ─────────
def main():
    # 라벨 로드
    y_true = pd.read_csv(DATA_CSV)["label"].values.astype(int)
    n = len(y_true)
    print(f"📂 {DATA_CSV}  ({n} samples, pos={int(y_true.sum())} / neg={int(n-y_true.sum())})")
    assert n == 16, f"기대 16, 실제 {n}"

    rows = []

    for ds in DATASETS:
        ds_pred_dir = f"{PRED_ROOT}/{ds}_newdata"
        ds_roc_dir  = f"{OUT_ROC_DIR}/{ds}"
        fold_preds  = []

        for k in range(1, 6):
            yp = np.load(f"{ds_pred_dir}/fold_{k}/predictions.npy").ravel()
            assert len(yp) == n, f"{ds} fold_{k}: shape mismatch ({len(yp)} vs {n})"
            m = compute_metrics(y_true, yp)
            row = {"Dataset": ds, "Fold": f"fold_{k}", "n_test": n, **m}
            rows.append(row)

            plot_roc_step(y_true, yp,
                          title=f"{ds} — fold {k}  (n=16)",
                          save_dir=ds_roc_dir, fname=f"fold_{k}",
                          color=CURVE_COLOR)
            print(f"  ✓ {ds}/fold_{k}.png  AUC={m['AUC-ROC']:.4f}")
            fold_preds.append(yp)

        # 5-fold avg
        avg = np.mean(fold_preds, axis=0)
        m   = compute_metrics(y_true, avg)
        rows.append({"Dataset": ds, "Fold": "5fold_avg", "n_test": n, **m})
        plot_roc_step(y_true, avg,
                      title=f"{ds} — 5-fold avg  (n=16)",
                      save_dir=ds_roc_dir, fname="5fold_avg",
                      color=AVG_COLOR)
        print(f"  ✓ {ds}/5fold_avg.png  AUC={m['AUC-ROC']:.4f}")

        # overlay
        folds_data = [(f"fold {k}", y_true, fold_preds[k-1], FOLD_COLORS[k-1])
                      for k in range(1, 6)]
        avg_data   = ("5-fold avg", y_true, avg, AVG_COLOR)
        plot_roc_overlay_step(folds_data, avg_data,
                              title=f"{ds} — DL ROC on 16 samples",
                              save_dir=ds_roc_dir, fname="overlay_5folds_avg")
        print(f"  ✓ {ds}/overlay_5folds_avg.png")

        # fold mean ± std (참고용)
        fold_only = [r for r in rows if r["Dataset"] == ds and r["Fold"].startswith("fold_")]
        df_f = pd.DataFrame(fold_only)
        rows.append({
            "Dataset":   ds,
            "Fold":      "fold_mean±std",
            "n_test":    n,
            "Accuracy":  f"{df_f['Accuracy'].mean():.4f}±{df_f['Accuracy'].std(ddof=1):.4f}",
            "Precision": f"{df_f['Precision'].mean():.4f}±{df_f['Precision'].std(ddof=1):.4f}",
            "Recall":    f"{df_f['Recall'].mean():.4f}±{df_f['Recall'].std(ddof=1):.4f}",
            "F1":        f"{df_f['F1'].mean():.4f}±{df_f['F1'].std(ddof=1):.4f}",
            "AUC-ROC":   f"{df_f['AUC-ROC'].mean():.4f}±{df_f['AUC-ROC'].std(ddof=1):.4f}",
            "AUC-PRC":   f"{df_f['AUC-PRC'].mean():.4f}±{df_f['AUC-PRC'].std(ddof=1):.4f}",
        })

    # ────── Excel 저장 (시트별 + 통합) ──────
    df_all = pd.DataFrame(rows)
    cols = ["Dataset", "Fold", "n_test",
            "Accuracy", "Precision", "Recall", "F1", "AUC-ROC", "AUC-PRC"]
    df_all = df_all[cols]

    # 숫자 4자리 포맷 (string은 그대로)
    def fmt(x):
        return x if isinstance(x, str) else f"{x:.4f}"
    metric_cols = ["Accuracy", "Precision", "Recall", "F1", "AUC-ROC", "AUC-PRC"]
    df_all_disp = df_all.copy()
    for c in metric_cols:
        df_all_disp[c] = df_all_disp[c].map(fmt)

    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as w:
        for ds in DATASETS:
            sub = df_all_disp[df_all_disp["Dataset"] == ds].drop(columns=["Dataset"])
            sub.to_excel(w, sheet_name=ds, index=False, startrow=2)
        df_all_disp.to_excel(w, sheet_name="all_combined", index=False, startrow=2)

    # ────── 서식 ──────
    wb = load_workbook(OUT_XLSX)
    HEAD_FILL = PatternFill(start_color="1A3D6E", end_color="1A3D6E", fill_type="solid")
    HEAD_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    TITLE_FONT = Font(name="Arial", size=14, bold=True, color="1A3D6E")
    SUB_FONT   = Font(name="Arial", size=10, italic=True, color="555555")
    BODY_FONT  = Font(name="Arial", size=10)
    AVG_FILL   = PatternFill(start_color="FFF2E0", end_color="FFF2E0", fill_type="solid")
    STAT_FILL  = PatternFill(start_color="E8F6E8", end_color="E8F6E8", fill_type="solid")
    BORDER = Border(left=Side(style="thin", color="CCCCCC"),
                    right=Side(style="thin", color="CCCCCC"),
                    top=Side(style="thin", color="CCCCCC"),
                    bottom=Side(style="thin", color="CCCCCC"))

    def style_sheet(ws, title, subtitle, n_cols, header_row=3):
        ws.cell(row=1, column=1, value=title).font = TITLE_FONT
        ws.cell(row=2, column=1, value=subtitle).font = SUB_FONT
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)

        for c in range(1, n_cols + 1):
            cell = ws.cell(row=header_row, column=c)
            cell.fill = HEAD_FILL
            cell.font = HEAD_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDER

        for r in range(header_row + 1, ws.max_row + 1):
            fold_cell_val = ws.cell(row=r, column=1).value or ""
            # all_combined 시트에서는 Fold 컬럼이 두 번째
            if "all_combined" in ws.title:
                fold_cell_val = ws.cell(row=r, column=2).value or ""
            for c in range(1, n_cols + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = BODY_FONT
                cell.border = BORDER
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if "5fold_avg" in fold_cell_val:
                for c in range(1, n_cols + 1):
                    ws.cell(row=r, column=c).fill = AVG_FILL
                    ws.cell(row=r, column=c).font = Font(name="Arial", size=10, bold=True)
            elif "fold_mean" in fold_cell_val:
                for c in range(1, n_cols + 1):
                    ws.cell(row=r, column=c).fill = STAT_FILL
                    ws.cell(row=r, column=c).font = Font(name="Arial", size=10, italic=True)

        widths = [12, 16, 8] + [12] * (n_cols - 3)
        for i, w in enumerate(widths[:n_cols], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="left")
        ws.cell(row=2, column=1).alignment = Alignment(horizontal="left")

    for ds in DATASETS:
        style_sheet(wb[ds],
                    title=f"DL on 16 samples — {ds}",
                    subtitle=f"PBertKla 5-fold predictions on new_data_v1.csv (n=16)  ·  threshold=0.5",
                    n_cols=8)
    style_sheet(wb["all_combined"],
                title="DL on 16 samples — All datasets",
                subtitle="PBertKla 5-fold predictions on new_data_v1.csv (n=16)  ·  threshold=0.5",
                n_cols=9)
    wb._sheets = [wb["data1"], wb["data2"], wb["data3"], wb["all_combined"]]
    wb.save(OUT_XLSX)

    # 콘솔 미리보기
    print()
    print("=" * 60)
    print("📊 DL on 16 samples — preview")
    print("=" * 60)
    print(df_all_disp.to_string(index=False))
    print(f"\n✅ Excel: {OUT_XLSX}  ({os.path.getsize(OUT_XLSX)/1024:.1f} KB)")
    print(f"✅ ROC dir: {OUT_ROC_DIR}/")


if __name__ == "__main__":
    main()
